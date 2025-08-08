import pdfplumber
import pandas as pd
import os
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SYMBOL_LEGEND = {
    "◎": "Chairperson",
    "●": "Secretariat",
    "○": "Member",
    "△": "Sub-member"
}

def extract_regular_table_from_page(page):
    """Extract regular table using enhanced pdfplumber logic for complex tables."""
    # Try different table extraction strategies
    tables = page.extract_tables()
    
    if not tables:
        # Try with more aggressive settings
        tables = page.extract_tables(
            table_settings={
                "vertical_strategy": "text",
                "horizontal_strategy": "text",
                "intersection_tolerance": 15,
                "edge_min_length": 3,
                "min_words_vertical": 1,
                "min_words_horizontal": 1
            }
        )
    
    if not tables:
        # Try with explicit table detection
        tables = page.find_tables()
        if tables:
            tables = [table.extract() for table in tables]

    extracted_tables = []

    for table in tables:
        if not table:
            continue
            
        # Convert to DataFrame
        df = pd.DataFrame(table)
        
        # Clean the data
        df = df.fillna('')
        
        # Remove completely empty rows and columns
        df.dropna(how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)
        
        # Clean cell values
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace('nan', '').replace('None', '')
        
        # Remove rows and columns that are completely empty after cleaning
        df = df.loc[:, ~df.apply(lambda col: col.astype(str).str.strip().eq('').all())]
        df = df.loc[~df.apply(lambda row: row.astype(str).str.strip().eq('').all(), axis=1)]
        
        if not df.empty and len(df) > 1:  # Ensure we have at least 2 rows
            extracted_tables.append(df)

    return extracted_tables

def clean_multiline_header(text):
    """Convert broken multiline text like 'C\nh\ni\ne\nf' into 'Chief'."""
    if isinstance(text, str):
        return ''.join(text.split()).strip()
    return text


def create_styled_excel(page4_tables, page5_tables):
    """Create styled Excel with proper committee structure for Page 4."""
    wb = Workbook()
    wb.remove(wb.active)

    # Style definitions
    title_font = Font(bold=True, size=16)
    section_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    executive_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    global_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Page 4 sheet with extracted data
    if page4_tables:
        df = page4_tables[0]  # Assume only 1 table on page 4
        ws = wb.create_sheet(title="Page4_Committee")

        # Clean the dataframe
        df = df.fillna('')
        
        # Get the actual data from the extracted table
        rows = df.values.tolist()
        rows = [[clean_multiline_header(cell) for cell in row] for row in rows]

        
        # Find the structure by analyzing the data
        # Look for main headers (EXECUTIVE OFFICERS, GLOBAL/REGIONAL DIRECTORS)
        executive_start_col = None
        global_start_col = None
        
        # Find the committee names (first column)
        committee_names = []
        for i, row in enumerate(rows):
            if i > 0 and row[0] and 'Committee' in str(row[0]):
                committee_names.append(str(row[0]).strip())
        
        # Find the main headers in the first row
        for i, cell in enumerate(rows[0]):
            cell_str = clean_multiline_header(cell).upper()

            if 'EXECUTIVE' in cell_str and 'OFFICER' in cell_str:
                executive_start_col = i
            elif 'GLOBAL' in cell_str or 'REGIONAL' in cell_str:
                global_start_col = i
        
        # If we can't find the headers, use default structure
        if executive_start_col is None:
            executive_start_col = 1
        if global_start_col is None:
            global_start_col = 8  # Default position after executive officers
        
        # Extract subheadings (second row)
        subheadings = []
        if len(rows) > 1:
            subheadings = [clean_multiline_header(cell) for cell in rows[1] if str(cell).strip()]

        
        # Create the structure
        # Row 1: Main headers
        if executive_start_col > 0:
            ws.merge_cells(start_row=1, start_column=executive_start_col + 1, 
                          end_row=1, end_column=global_start_col)
            ws.cell(row=1, column=executive_start_col + 1).value = "EXECUTIVE OFFICERS"
            ws.cell(row=1, column=executive_start_col + 1).font = title_font
            ws.cell(row=1, column=executive_start_col + 1).alignment = center_align
            ws.cell(row=1, column=executive_start_col + 1).fill = executive_fill

        if global_start_col < len(rows[0]):
            ws.merge_cells(start_row=1, start_column=global_start_col + 1, 
                          end_row=1, end_column=len(rows[0]))
            ws.cell(row=1, column=global_start_col + 1).value = "GLOBAL/REGIONAL DIRECTORS"
            ws.cell(row=1, column=global_start_col + 1).font = title_font
            ws.cell(row=1, column=global_start_col + 1).alignment = center_align
            ws.cell(row=1, column=global_start_col + 1).fill = global_fill

        # Row 1, Column 1: "Committees" header
        ws.cell(row=1, column=1).value = "Committees"
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).fill = header_fill

        # Row 2: Subheadings (shifted one column to the right)
        for i, cell in enumerate(rows[1]):
            if str(cell).strip():
                col_num = i   # Shift headings one column to the right
                if i >= executive_start_col and i < global_start_col:
                    fill_color = executive_fill
                elif i >= global_start_col:
                    fill_color = global_fill
                else:
                    fill_color = header_fill
                
                ws.cell(row=2, column=col_num).value = str(cell).strip()
                ws.cell(row=2, column=col_num).font = subheader_font
                ws.cell(row=2, column=col_num).alignment = center_align
                ws.cell(row=2, column=col_num).fill = fill_color

        # Rows 3+: Committee names and symbols
        for row_idx, row in enumerate(rows[2:], start=3):
            for col_idx, cell in enumerate(row):
                if str(cell).strip():
                    ws.cell(row=row_idx, column=col_idx + 1).value = str(cell).strip()
                    
                    # Apply styling based on column position
                    if col_idx == 0:  # Committee names column
                        ws.cell(row=row_idx, column=col_idx + 1).font = subheader_font
                        ws.cell(row=row_idx, column=col_idx + 1).alignment = left_align
                        ws.cell(row=row_idx, column=col_idx + 1).fill = header_fill
                    else:  # Symbol columns
                        ws.cell(row=row_idx, column=col_idx + 1).alignment = center_align

        # Set column widths
        ws.column_dimensions['A'].width = 25
        from openpyxl.utils import get_column_letter
        for col in range(2, len(rows[0]) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15

    # Page 5 regular tables
    if page5_tables:
        for i, df in enumerate(page5_tables):
            ws = wb.create_sheet(title=f"Page5_Table{i+1}"[:31])
            df = df.fillna('').applymap(clean_multiline_header)

            for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
                cleaned_row = [clean_multiline_header(cell) for cell in r]
                ws.append(cleaned_row)

            # Optional: Set width and alignment
            for col in range(1, df.shape[1] + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
                for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                    for cell in row:
                        cell.alignment = center_align

    output_excel = "final_solution_output.xlsx"
    wb.save(output_excel)
    return wb, output_excel


def extract_pages_4_and_5(pdf_path):
    """Extract both pages using same logic."""
    print(f"Starting extraction from: {pdf_path}")
    page4_tables, page5_tables = [], []

    with pdfplumber.open(pdf_path) as pdf:
        if len(pdf.pages) < 5:
            print("PDF must have at least 5 pages.")
            return [], []

        print("Processing Page 4...")
        page4 = pdf.pages[3]
        page4_tables = extract_regular_table_from_page(page4)

        print("Processing Page 5...")
        page5 = pdf.pages[4]
        page5_tables = extract_regular_table_from_page(page5)

    return page4_tables, page5_tables

def clean_multiline_header(text):
    """Convert broken multiline text like 'C\nh\ni\ne\nf' into 'Chief'."""
    if isinstance(text, str):
        return ''.join(text.split()).strip()
    return text

def create_committee_json_from_excel(excel_path, json_path="committee_structure.json"):
    df = pd.read_excel(excel_path, sheet_name="Page4_Committee", header=[0, 1])
    df.columns = [(clean_multiline_header(a), clean_multiline_header(b)) for a, b in df.columns]
    df.rename(columns={df.columns[0]: ("COMMITTEE", "NAME")}, inplace=True)
    df.dropna(axis=1, how='all', inplace=True)
    df.set_index(("COMMITTEE", "NAME"), inplace=True)

    nested_json = {}
    for committee_name in df.index:
        committee_entry = {
            "EXECUTIVE OFFICERS": OrderedDict(),
            "GLOBAL/REGIONAL DIRECTORS": OrderedDict()
        }
        for (section, role), value in df.loc[committee_name].items():
            symbol = str(value).strip()
            if not symbol or symbol.lower() == 'nan':
                continue

            legend = SYMBOL_LEGEND.get(symbol, "Unknown")
            formatted_value = f"{symbol} ({legend})"

            if "EXECUTIVE" in section.upper():
                committee_entry["EXECUTIVE OFFICERS"][role] = formatted_value
            elif "GLOBAL" in section.upper() or "REGIONAL" in section.upper():
                committee_entry["GLOBAL/REGIONAL DIRECTORS"][role] = formatted_value
            else:
                committee_entry["EXECUTIVE OFFICERS"].setdefault(role, formatted_value)

        nested_json[committee_name] = committee_entry

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(nested_json, f, indent=4, ensure_ascii=False)

    print(f"✅ Cleaned JSON saved to: {json_path}")
    return nested_json

def create_combined_json_from_excel(excel_path, json_path="combined_output.json"):
    wb = load_workbook(excel_path, data_only=True)
    combined_data = {}
    print("🔍 Processing 'Page4_Committee'...")
    page4_json = create_committee_json_from_excel(excel_path)
    combined_data["Page4_Committee"] = page4_json

    print("🔍 Processing 'Page5_' sheets...")
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Page5_"):
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            df = df.fillna('').applymap(clean_multiline_header)
            headers = df.iloc[1].tolist()
            headers = [clean_multiline_header(h) for h in headers]
            df = df.iloc[2:].reset_index(drop=True)

            table_data = OrderedDict()
            for _, row in df.iterrows():
                committee_name = str(row.iloc[0]).strip()
                if not committee_name:
                    continue

                entry = OrderedDict()
                for i in range(1, len(row)):
                    value = str(row.iloc[i]).strip()
                    if value and value not in ['nan', 'None']:
                        header = headers[i]
                        roles = []
                        for symbol in value:
                            if symbol in SYMBOL_LEGEND:
                                roles.append(f"{symbol} ({SYMBOL_LEGEND[symbol]})")
                        if roles:
                            entry[header] = roles

                if entry:
                    table_data[committee_name] = entry

            combined_data[sheet_name] = table_data

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(combined_data, f, indent=4, ensure_ascii=False)

    print(f"✅ Combined JSON saved to: {json_path}")
    return combined_data

def main():
    pdf_path = "new_pdf\\16 Committee Regulation.pdf"
    page4_tables, page5_tables = extract_pages_4_and_5(pdf_path)

    if not page4_tables and not page5_tables:
        print("No data extracted.")
        return

    print("Creating Excel output...")
    wb, output_excel = create_styled_excel(page4_tables, page5_tables)

    print("Creating JSON from Excel structure...")
    combined_json = create_combined_json_from_excel(output_excel)

    print("Extraction complete.")
    print(f"Excel saved to: {output_excel}")
    print(f"JSON structure saved and ready.")

if __name__ == "__main__":
    main()
