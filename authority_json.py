import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re
import json

# === SETTINGS ===
file_path = "pdf/15 Official Authority Regulations_20250425.pdf"
output_excel = "authority.xlsx"
target_pages = [10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28]

# Define columns for output
columns = [
    "S/N", "Classification","Sub_Classification", "MOL", "BDM", "A1", "A2", "A3", "A4", "A5",
    "Co-Mgmt. Dept.", "Deliberation MM", "Report MM", "Report A3", "Review GPM", "CC Dept."
]

footer_notes = []  # ⬅️ store footer notes globally


# Merged header definitions: (text, start_col_idx, end_col_idx)
merge_headers = [
    ("S/N", 1, 1),
    ("Classification", 2, 3),
    ("Authorized Approver", 4, 10),
    ("Co-Mgmt.", 11, 11),
    ("Deliberation", 12, 12),
    ("Report", 12, 14),
    ("Review", 15, 15),
    ("CC", 16, 16)
]

# === HELPER FUNCTIONS ===
def clean_text(text):
    """Clean and normalize text data"""
    if pd.isna(text):
        return ""
    text = str(text).strip()
    text = re.sub(r'\s+', ' ', text)  # Replace multiple spaces with single space
    return text

def contains_only_dots(text):
    """Return True only if the text is only ● or bullet-like symbols, not real content."""
    if not isinstance(text, str):
        return False
    text = text.strip()
    return text in {"●", "•", "▪", "–", "-", "‣"}

def is_footer_note_row(row_data):
    """Check if a row contains footer notes/explanations that should be excluded"""
    if not row_data or len(row_data) == 0:
        return False
    
    # Convert row to string for analysis
    row_text = " ".join([str(cell) if cell is not None else "" for cell in row_data]).strip()
    
    # Skip empty rows
    if not row_text or row_text.replace("None", "").strip() == "":
        return True
    
    # Patterns that indicate footer notes/explanations
    footer_patterns = [
        r"Responsible Department",
        r"Conclusion/\s*Termination/\s*Revision of Service Agreement",
        r"Commercial Subsidiaries.*GPM Dept",
        r"Ship Management Subsidiaries.*Ship Management Dept",
        r"Decision/\s*Payment of Service Fee",
        r"GAF Dept\.",
        r"\*Approval is not required",
        r"Only applicable for intergroup service agreements",
        r"Refer to.*Regulations Handling Regulations",
        r"No approval is required for amendments",
        r"No approval is required for the logical amendments"
    ]
    
    # Check if row matches any footer pattern
    for pattern in footer_patterns:
        if re.search(pattern, row_text, re.IGNORECASE):
            return True
    
    # Check if row starts with explanatory text (not table content)
    # Table content typically starts with S/N numbers or specific classifications
    first_cell = str(row_data[0] if row_data[0] is not None else "").strip()
    
    # If first cell contains long explanatory text, it's likely a footer note
    if len(first_cell) > 50 and not re.match(r'^\(?[①-⑳0-9]{1,3}\)?', first_cell):
        return True
    
    return False

def shift_row_right(row, df_columns, shift_from_col="Sub_Classification"):
    """Shift row data one column to the right starting from specified column"""
    if shift_from_col not in df_columns:
        print(f"⚠️ Column '{shift_from_col}' not found in dataframe")
        return row
    
    start_idx = df_columns.get_loc(shift_from_col)
    new_row = row.copy()
    
    # Shift values to the right starting from Sub_Classification
    # Go backwards through columns to avoid overwriting
    for i in range(len(df_columns) - 1, start_idx, -1):
        col_current = df_columns[i]
        col_previous = df_columns[i - 1]
        old_val = new_row.get(col_previous, pd.NA)
        new_row[col_current] = old_val
        print(f"   📤 {col_previous} → {col_current}: '{old_val}'")
    
    # Make the starting column (Sub_Classification) empty
    new_row[shift_from_col] = pd.NA
    
    return new_row

def is_currency(text):
    """Check if text contains currency symbols"""
    if not isinstance(text, str):
        return False
    return any(symbol in text for symbol in ["¥", "$", "US$"])

def extract_sn(text):
    """Extract S/N from classification text if it starts with a number"""
    if not isinstance(text, str):
        return None, text
    match = re.match(r'^(\d+)(?:\s+|\.|\))?(.*)', text.strip())
    if match:
        return match.group(1), match.group(2).strip()
    return None, text

def validate_table(df):
    """Validate and clean the dataframe"""
    # Ensure we have enough columns
    if len(df.columns) < 2:  # At least S/N and Classification
        return None
    
    # Clean all data
    df = df.map(clean_text)
    
    # Replace empty strings with NaN for consistency
    df.replace("", pd.NA, inplace=True)
    
    return df

def is_valid_sn(sn: str) -> bool:
    """Check if the S/N value is a valid entry like (1), ①, ②, etc."""
    sn = sn.strip()
    return bool(re.match(r"^\(?[①-⑳0-9]{1,3}\)?$", sn))  # Matches (1), ①, 1, etc.

def filter_table_content(table_data):
    """Filter out header/placeholder rows and footer notes"""
    if not table_data:
        return []

    filtered_rows = []

    for row in table_data:
        # Convert to strings
        row_text = [str(cell or "").strip() for cell in row]
        row_joined = " ".join(row_text)

        # Skip footer note rows
        if is_footer_note_row(row):
            note = row_joined.strip()   
            if note:
                footer_notes.append(note)
                print(f"📝 Collected footer note: {note[:80]}")
            continue

        # Skip known header-like rows (e.g., contain MOL, A1, BDM...)
        if all(
            item in {"MOL", "BDM", "A1", "A2", "A3", "A4", "A5", "Dept.", "MM", "GPM", "A3"}
            for item in row_text if item
        ):
            print(f"⏭️ Skipping header-like row: {row_text}")
            continue

        filtered_rows.append(row)

    return filtered_rows


# === MAIN SCRIPT ===
all_tables = []

with pdfplumber.open(file_path) as pdf:
    for page_num in target_pages:
        try:
            if page_num - 1 >= len(pdf.pages):
                continue
                
            page = pdf.pages[page_num - 1]
            tables = page.extract_tables()

            # Extract possible footers from full page text (outside of table)
            text_lines = page.extract_text().split('\n')
            for line in text_lines:
                line = line.strip()
                if re.match(r"^\d{1,2}\s", line):  # Line starts with number and space
                    footer_notes.append(line)
                    print(f"📥 Found footer line on page {page_num}: {line}")
            
            if not tables:
                print(f"⚠️ No tables found on page {page_num}")
                continue
                
            for table in tables:
                # Filter out footer notes BEFORE creating DataFrame
                filtered_table = filter_table_content(table)
                
                if not filtered_table:
                    print(f"⚠️ No valid table content after filtering on page {page_num}")
                    continue
                
                df = pd.DataFrame(filtered_table)
                
                # Basic cleaning
                df.dropna(how='all', inplace=True)
                df.dropna(axis=1, how='all', inplace=True)
                df.reset_index(drop=True, inplace=True)
                
                # Validate table structure
                df = validate_table(df)
                if df is None:
                    print(f"⚠️ Invalid table structure on page {page_num}")
                    continue
                
                # Use first row as header if it looks like headers
                if (df.shape[0] > 1 and 
                    not df.iloc[0].isnull().all() and 
                    len(df.iloc[0]) == len(df.columns)):
                    potential_header = df.iloc[0].apply(clean_text)
                    if any(len(x) > 0 for x in potential_header):
                        df.columns = potential_header
                        df = df[1:].reset_index(drop=True)
                
                # Rename columns to our fixed schema
                available_cols = min(len(df.columns), len(columns))
                df.columns = columns[:available_cols]
                
                # Add any missing columns
                for col in columns[available_cols:]:
                    df[col] = pd.NA
                
                # === DATA CLEANING ===
                for idx, row in df.iterrows():
                    # Fix currency text in wrong columns
                    for col in ["MOL", "BDM", "A1", "A2", "A3", "A4", "A5"]:
                        if col in df.columns:
                            val = row.get(col, "")
                            if is_currency(val):
                                old_class = str(row.get("Classification", ""))
                                df.at[idx, "Classification"] = f"{val.strip()} {old_class}".strip()
                                df.at[idx, col] = pd.NA
                    
                    # Fix S/N text in Classification column
                    if "S/N" in df.columns and "Classification" in df.columns:
                        sn_val = row.get("S/N", "")
                        class_val = row.get("Classification", "")
                        
                        if pd.isna(sn_val) or str(sn_val).strip() == "":
                            sn, new_class = extract_sn(class_val)
                            if sn:
                                df.at[idx, "S/N"] = sn
                                df.at[idx, "Classification"] = new_class

                # === DOT SHIFTING LOGIC ===
                if "Sub_Classification" in df.columns:
                    sub_col_index = df.columns.get_loc("Sub_Classification")

                    # Check if ANY row has a dot in Sub_Classification
                    dot_in_any = df["Sub_Classification"].astype(str).apply(str.strip).apply(contains_only_dots).any()

                    if dot_in_any:
                        print("🔁 Dot found in Sub_Classification → Shifting entire table right from this column.")

                        # Shift each row's values from Sub_Classification onwards by 1 to the right
                        for idx in df.index:
                            for i in range(len(df.columns) - 1, sub_col_index, -1):
                                df.iat[idx, i] = df.iat[idx, i - 1]
                            # Clear the original dot in Sub_Classification column
                            df.iat[idx, sub_col_index] = pd.NA
                
                # Final check to ensure we have actual table content
                if df.empty or df.shape[0] == 0:
                    print(f"⚠️ No data rows remaining after cleaning on page {page_num}")
                    continue
                
                print(f"✅ Processed table from page {page_num} with {df.shape[0]} rows and {df.shape[1]} columns")
                all_tables.append((page_num, df))
                
        except Exception as e:
            print(f"⚠️ Error processing page {page_num}: {str(e)}")
            continue

# === EXPORT TO EXCEL ===
if not all_tables:
    print("⚠️ No valid tables found to export")
else:
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for i, (page_num, df) in enumerate(all_tables, 1):
            sheet_name = f"Page{page_num}_Table{i}"
            
            # Write data (starting at row 3 to leave room for headers)
            df.to_excel(
                writer, 
                sheet_name=sheet_name, 
                index=False, 
                header=False, 
                startrow=2
            )

    # === POST-PROCESSING: Formatting Headers ===
    try:
        wb = load_workbook(output_excel)

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            # Write second-row subheaders (row 3)
            for col_num, col_name in enumerate(columns, 1):
                if col_num <= ws.max_column:  # Only write if column exists
                    ws.cell(row=3, column=col_num, value=col_name)
                    ws.cell(row=3, column=col_num).alignment = Alignment(
                        horizontal="center", 
                        vertical="center"
                    )

            # Write top-row merged headers (row 2)
            for heading, start_col, end_col in merge_headers:
                if start_col <= ws.max_column:  # Only write if column exists
                    cell = ws.cell(row=2, column=start_col)
                    cell.value = heading
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if start_col != end_col and end_col <= ws.max_column:
                        start_letter = get_column_letter(start_col)
                        end_letter = get_column_letter(end_col)
                        ws.merge_cells(f"{start_letter}2:{end_letter}2")

            # Adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Check header and first few rows for max length
                cells_to_check = list(column[:10])  # First 10 rows
                if len(column) > 10:
                    cells_to_check.extend(column[-5:])  # Last 5 rows
                
                for cell in cells_to_check:
                    try:
                        if cell.value:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

            # Adjust row heights
            ws.row_dimensions[2].height = 20    
            ws.row_dimensions[3].height = 20

        wb.save(output_excel)
        print(f"✅ Excel saved to: {output_excel}")
        
    except Exception as e:
        print(f"⚠️ Error during Excel formatting: {str(e)}")

        

# === EXPORT TO JSON (Structured Format) ===
def create_nested_structure(tables):
    """Convert flat tables into nested structure based on S/N and Sub_Classification"""
    structured_data = {}

    for i, (page_num, df) in enumerate(tables, 1):
        table_key = f"Page{page_num}_Table{i}"
        structured_data[table_key] = []

        current_entry = None
        last_classification = ""
        last_sn = ""

        for _, row in df.iterrows():
            sn = clean_text(row.get("S/N", ""))
            classification = clean_text(row.get("Classification", ""))
            sub_classification = clean_text(row.get("Sub_Classification", ""))

            # Update and inherit classification
            if classification:
                last_classification = classification
            else:
                classification = last_classification

            # Prepare row_data excluding S/N and Classification fields
            row_data = {
                col: clean_text(row[col])
                for col in df.columns
                if col not in ["S/N", "Classification", "Sub_Classification"]
            }

            # New block if a valid S/N is found
            if sn and is_valid_sn(sn):
                last_sn = sn
                if current_entry:
                    structured_data[table_key].append(current_entry)

                current_entry = {
                    "S/N": sn,
                    "Classification": classification,
                    "Sub_Items": []
                }

            if current_entry is None:
                # Initialize current_entry if first rows are missing S/N
                current_entry = {
                    "S/N": last_sn or "",
                    "Classification": classification,
                    "Sub_Items": []
                }

            sub_item = {
                "Sub_Classification": sub_classification or classification,
                **row_data
            }

            # Only add sub-item if any real data present
            if any(v for v in sub_item.values()):
                current_entry["Sub_Items"].append(sub_item)

        # Final entry append
        if current_entry:
            structured_data[table_key].append(current_entry)

    return structured_data



# Create structured data
structured_tables = create_nested_structure(all_tables)

# Save to JSON
json_path = "authority_json.json"
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(structured_tables, f, ensure_ascii=False, indent=2)

print(f"✅ Structured tables saved to: {json_path}")
print(f"📊 Total tables processed: {len(all_tables)}")


if footer_notes:
    with open("footer_notes.json", "w", encoding="utf-8") as f:
        json.dump(footer_notes, f, indent=2, ensure_ascii=False)
    print(f"✅ Saved {len(footer_notes)} footer notes to 'footer_notes.json'")
else:
    print("ℹ️ No footer notes found.")